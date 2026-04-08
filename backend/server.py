from fastapi import FastAPI, APIRouter, HTTPException, Depends, status
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from fastapi.responses import StreamingResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict, EmailStr
from typing import List, Optional
import uuid
from datetime import datetime, timezone, timedelta
import jwt
import bcrypt
import io

from apscheduler.schedulers.asyncio import AsyncIOScheduler
from apscheduler.triggers.cron import CronTrigger

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

try:
    from twilio.rest import Client as TwilioClient
    TWILIO_AVAILABLE = True
except ImportError:
    TWILIO_AVAILABLE = False

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

JWT_SECRET = os.environ.get('JWT_SECRET', 'your-secret-key-change-in-production')
JWT_ALGORITHM = 'HS256'

TWILIO_ACCOUNT_SID  = os.environ.get('TWILIO_ACCOUNT_SID', '')
TWILIO_AUTH_TOKEN   = os.environ.get('TWILIO_AUTH_TOKEN', '')
TWILIO_WHATSAPP_FROM = os.environ.get('TWILIO_WHATSAPP_FROM', 'whatsapp:+14155238886')

app = FastAPI()
api_router = APIRouter(prefix="/api")
security = HTTPBearer()
scheduler = AsyncIOScheduler()

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# ============ MODELS ============

class UserRegister(BaseModel):
    email: EmailStr
    password: str
    nombre: str

class UserLogin(BaseModel):
    email: EmailStr
    password: str

class User(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    email: str
    nombre: str

class TokenResponse(BaseModel):
    token: str
    user: User

class SocioCreate(BaseModel):
    nombre: str
    email: Optional[str] = None
    telefono: Optional[str] = None
    direccion: Optional[str] = None

class Socio(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    socio_id: str
    nombre: str
    email: Optional[str] = None
    telefono: Optional[str] = None
    direccion: Optional[str] = None
    fecha_registro: str
    estado: str
    fecha_vencimiento: Optional[str] = None

class PagoCreate(BaseModel):
    socio_id: str
    monto: float
    tipo_plan: str
    metodo_pago: Optional[str] = "Efectivo"

class Pago(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    socio_id: str
    monto: float
    tipo_plan: str
    metodo_pago: str
    fecha_pago: str
    fecha_vencimiento: str

class DashboardStats(BaseModel):
    total_socios: int
    socios_activos: int
    socios_vencidos: int
    ingresos_mes: float
    proximos_vencimientos: List[dict]

class AlertaResult(BaseModel):
    enviadas: int
    errores: int
    sin_telefono: int
    detalle: List[dict]

# ============ AUTH HELPERS ============

def hash_password(password: str) -> str:
    salt = bcrypt.gensalt()
    return bcrypt.hashpw(password.encode('utf-8'), salt).decode('utf-8')

def verify_password(password: str, hashed: str) -> bool:
    return bcrypt.checkpw(password.encode('utf-8'), hashed.encode('utf-8'))

def create_token(user_id: str) -> str:
    payload = {'user_id': user_id, 'exp': datetime.now(timezone.utc) + timedelta(days=7)}
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGORITHM)

async def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)):
    try:
        token = credentials.credentials
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        user_id = payload.get('user_id')
        user = await db.users.find_one({'id': user_id}, {'_id': 0, 'password': 0})
        if not user:
            raise HTTPException(status_code=401, detail="Usuario no encontrado")
        return user
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expirado")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Token inválido")

# ============ HELPER FUNCTIONS ============

def calcular_fecha_vencimiento(tipo_plan: str, fecha_base: datetime) -> datetime:
    dias = {'mensual': 30, 'trimestral': 90, 'semestral': 180, 'anual': 365}
    return fecha_base + timedelta(days=dias.get(tipo_plan, 30))

async def actualizar_estado_socio(socio_id: str):
    ultimo_pago = await db.pagos.find_one({'socio_id': socio_id}, {'_id': 0}, sort=[('fecha_vencimiento', -1)])
    if ultimo_pago:
        fecha_venc = datetime.fromisoformat(ultimo_pago['fecha_vencimiento'])
        estado = 'activo' if fecha_venc >= datetime.now(timezone.utc) else 'vencido'
        await db.socios.update_one({'socio_id': socio_id}, {'$set': {'estado': estado, 'fecha_vencimiento': ultimo_pago['fecha_vencimiento']}})
    else:
        await db.socios.update_one({'socio_id': socio_id}, {'$set': {'estado': 'vencido', 'fecha_vencimiento': None}})

# ============ WHATSAPP / ALERTAS ============

def enviar_whatsapp(telefono: str, mensaje: str) -> bool:
    if not TWILIO_AVAILABLE:
        logger.warning("Twilio no instalado. Ejecutar: pip install twilio")
        return False
    if not TWILIO_ACCOUNT_SID or not TWILIO_AUTH_TOKEN:
        logger.warning("Variables TWILIO_ACCOUNT_SID / TWILIO_AUTH_TOKEN no configuradas")
        return False
    try:
        twilio_client = TwilioClient(TWILIO_ACCOUNT_SID, TWILIO_AUTH_TOKEN)
        numero = telefono.strip()
        if not numero.startswith('+'):
            numero = '+' + numero
        twilio_client.messages.create(from_=TWILIO_WHATSAPP_FROM, to=f'whatsapp:{numero}', body=mensaje)
        return True
    except Exception as e:
        logger.error(f"Error enviando WhatsApp a {telefono}: {e}")
        return False

async def ejecutar_alertas_diarias():
    logger.info("Ejecutando alertas diarias de WhatsApp...")
    now = datetime.now(timezone.utc)
    socios = await db.socios.find({}, {'_id': 0}).to_list(10000)

    enviadas = errores = sin_telefono = 0

    for socio in socios:
        if not socio.get('fecha_vencimiento'):
            continue

        fecha_venc = datetime.fromisoformat(socio['fecha_vencimiento'])
        dias = (fecha_venc - now).days
        telefono = socio.get('telefono', '').strip()

        if dias < 0:
            mensaje = (
                f"Hola {socio['nombre']} 👋\n"
                f"Te recordamos que tu cuota en el gimnasio *venció* el {fecha_venc.strftime('%d/%m/%Y')}.\n"
                f"¡Acercate a renovarla para seguir entrenando! 💪"
            )
        elif 0 <= dias <= 7:
            if dias == 0:
                aviso = "vence *hoy*"
            elif dias == 1:
                aviso = "vence *mañana*"
            else:
                aviso = f"vence en *{dias} días*"
            mensaje = (
                f"Hola {socio['nombre']} 👋\n"
                f"Te avisamos que tu cuota del gimnasio {aviso} ({fecha_venc.strftime('%d/%m/%Y')}).\n"
                f"¡No te olvides de renovarla! 💪"
            )
        else:
            continue

        alerta_id = str(uuid.uuid4())
        await db.alertas.insert_one({
            'id': alerta_id,
            'socio_id': socio['socio_id'],
            'nombre': socio['nombre'],
            'tipo': 'vencido' if dias < 0 else 'proximo',
            'dias_restantes': dias,
            'fecha_vencimiento': socio['fecha_vencimiento'],
            'telefono': telefono,
            'mensaje': mensaje,
            'enviado': False,
            'fecha_alerta': now.isoformat()
        })

        if not telefono:
            sin_telefono += 1
            continue

        ok = enviar_whatsapp(telefono, mensaje)
        if ok:
            enviadas += 1
            await db.alertas.update_one({'id': alerta_id}, {'$set': {'enviado': True}})
        else:
            errores += 1

    logger.info(f"Alertas: {enviadas} enviadas, {errores} errores, {sin_telefono} sin teléfono")

# ============ AUTH ROUTES ============

@api_router.post("/auth/register", response_model=TokenResponse)
async def register(user_data: UserRegister):
    existing = await db.users.find_one({'email': user_data.email}, {'_id': 0})
    if existing:
        raise HTTPException(status_code=400, detail="El email ya está registrado")
    user_id = str(uuid.uuid4())
    hashed_pw = hash_password(user_data.password)
    user_doc = {'id': user_id, 'email': user_data.email, 'password': hashed_pw, 'nombre': user_data.nombre, 'created_at': datetime.now(timezone.utc).isoformat()}
    await db.users.insert_one(user_doc)
    token = create_token(user_id)
    return TokenResponse(token=token, user=User(id=user_id, email=user_data.email, nombre=user_data.nombre))

@api_router.post("/auth/login", response_model=TokenResponse)
async def login(credentials: UserLogin):
    user = await db.users.find_one({'email': credentials.email}, {'_id': 0})
    if not user or not verify_password(credentials.password, user['password']):
        raise HTTPException(status_code=401, detail="Credenciales inválidas")
    token = create_token(user['id'])
    return TokenResponse(token=token, user=User(id=user['id'], email=user['email'], nombre=user['nombre']))

@api_router.get("/auth/me", response_model=User)
async def get_me(current_user: dict = Depends(get_current_user)):
    return User(**current_user)

# ============ SOCIOS ROUTES ============

@api_router.post("/socios", response_model=Socio)
async def crear_socio(socio_data: SocioCreate, current_user: dict = Depends(get_current_user)):
    count = await db.socios.count_documents({})
    socio_id = f"GYM-{count + 1:04d}"
    socio_doc = {'id': str(uuid.uuid4()), 'socio_id': socio_id, 'nombre': socio_data.nombre, 'email': socio_data.email, 'telefono': socio_data.telefono, 'direccion': socio_data.direccion, 'fecha_registro': datetime.now(timezone.utc).isoformat(), 'estado': 'vencido', 'fecha_vencimiento': None}
    await db.socios.insert_one(socio_doc)
    socio_doc.pop('_id', None)
    return Socio(**socio_doc)

@api_router.get("/socios", response_model=List[Socio])
async def listar_socios(current_user: dict = Depends(get_current_user)):
    socios = await db.socios.find({}, {'_id': 0}).sort('fecha_registro', -1).to_list(1000)
    return [Socio(**s) for s in socios]

@api_router.get("/socios/{socio_id}", response_model=Socio)
async def obtener_socio(socio_id: str, current_user: dict = Depends(get_current_user)):
    socio = await db.socios.find_one({'socio_id': socio_id}, {'_id': 0})
    if not socio:
        raise HTTPException(status_code=404, detail="Socio no encontrado")
    return Socio(**socio)

@api_router.put("/socios/{socio_id}", response_model=Socio)
async def actualizar_socio(socio_id: str, socio_data: SocioCreate, current_user: dict = Depends(get_current_user)):
    result = await db.socios.update_one({'socio_id': socio_id}, {'$set': socio_data.model_dump()})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Socio no encontrado")
    socio = await db.socios.find_one({'socio_id': socio_id}, {'_id': 0})
    return Socio(**socio)

@api_router.delete("/socios/{socio_id}")
async def eliminar_socio(socio_id: str, current_user: dict = Depends(get_current_user)):
    result = await db.socios.delete_one({'socio_id': socio_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Socio no encontrado")
    await db.pagos.delete_many({'socio_id': socio_id})
    return {'message': 'Socio eliminado correctamente'}

# ============ PAGOS ROUTES ============

@api_router.post("/pagos", response_model=Pago)
async def registrar_pago(pago_data: PagoCreate, current_user: dict = Depends(get_current_user)):
    socio = await db.socios.find_one({'socio_id': pago_data.socio_id}, {'_id': 0})
    if not socio:
        raise HTTPException(status_code=404, detail="Socio no encontrado")
    fecha_pago = datetime.now(timezone.utc)
    fecha_vencimiento = calcular_fecha_vencimiento(pago_data.tipo_plan, fecha_pago)
    pago_doc = {'id': str(uuid.uuid4()), 'socio_id': pago_data.socio_id, 'monto': pago_data.monto, 'tipo_plan': pago_data.tipo_plan, 'metodo_pago': pago_data.metodo_pago, 'fecha_pago': fecha_pago.isoformat(), 'fecha_vencimiento': fecha_vencimiento.isoformat()}
    await db.pagos.insert_one(pago_doc)
    await actualizar_estado_socio(pago_data.socio_id)
    pago_doc.pop('_id', None)
    return Pago(**pago_doc)

@api_router.get("/pagos", response_model=List[Pago])
async def listar_pagos(current_user: dict = Depends(get_current_user)):
    pagos = await db.pagos.find({}, {'_id': 0}).sort('fecha_pago', -1).to_list(1000)
    return [Pago(**p) for p in pagos]

@api_router.get("/pagos/socio/{socio_id}", response_model=List[Pago])
async def obtener_pagos_socio(socio_id: str, current_user: dict = Depends(get_current_user)):
    pagos = await db.pagos.find({'socio_id': socio_id}, {'_id': 0}).sort('fecha_pago', -1).to_list(1000)
    return [Pago(**p) for p in pagos]

# ============ DASHBOARD ROUTES ============

@api_router.get("/dashboard/stats", response_model=DashboardStats)
async def obtener_stats(current_user: dict = Depends(get_current_user)):
    socios = await db.socios.find({}, {'_id': 0}).to_list(10000)
    for socio in socios:
        await actualizar_estado_socio(socio['socio_id'])
    socios = await db.socios.find({}, {'_id': 0}).to_list(10000)
    total_socios = len(socios)
    socios_activos = len([s for s in socios if s.get('estado') == 'activo'])
    socios_vencidos = len([s for s in socios if s.get('estado') == 'vencido'])
    now = datetime.now(timezone.utc)
    inicio_mes = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    pagos_mes = await db.pagos.find({'fecha_pago': {'$gte': inicio_mes.isoformat()}}, {'_id': 0}).to_list(10000)
    ingresos_mes = sum(p['monto'] for p in pagos_mes)
    proximos = []
    for socio in socios:
        if socio.get('fecha_vencimiento'):
            fecha_venc = datetime.fromisoformat(socio['fecha_vencimiento'])
            dias_restantes = (fecha_venc - now).days
            if 0 <= dias_restantes <= 7:
                proximos.append({'socio_id': socio['socio_id'], 'nombre': socio['nombre'], 'fecha_vencimiento': socio['fecha_vencimiento'], 'dias_restantes': dias_restantes})
    proximos.sort(key=lambda x: x['dias_restantes'])
    return DashboardStats(total_socios=total_socios, socios_activos=socios_activos, socios_vencidos=socios_vencidos, ingresos_mes=ingresos_mes, proximos_vencimientos=proximos[:10])

# ============ ALERTAS ROUTES ============

@api_router.get("/alertas/estado")
async def estado_alertas(current_user: dict = Depends(get_current_user)):
    now = datetime.now(timezone.utc)
    socios = await db.socios.find({}, {'_id': 0}).to_list(10000)
    proximos, vencidos = [], []
    for socio in socios:
        if not socio.get('fecha_vencimiento'):
            continue
        fecha_venc = datetime.fromisoformat(socio['fecha_vencimiento'])
        dias = (fecha_venc - now).days
        entry = {'socio_id': socio['socio_id'], 'nombre': socio['nombre'], 'telefono': socio.get('telefono', '') or '', 'fecha_vencimiento': socio['fecha_vencimiento'], 'dias_restantes': dias}
        if dias < 0:
            vencidos.append(entry)
        elif dias <= 7:
            proximos.append(entry)
    return {
        'proximos_a_vencer': sorted(proximos, key=lambda x: x['dias_restantes']),
        'vencidos': sorted(vencidos, key=lambda x: x['dias_restantes']),
        'twilio_configurado': bool(TWILIO_ACCOUNT_SID and TWILIO_AUTH_TOKEN)
    }

@api_router.post("/alertas/enviar-ahora", response_model=AlertaResult)
async def enviar_alertas_ahora(current_user: dict = Depends(get_current_user)):
    await ejecutar_alertas_diarias()
    now = datetime.now(timezone.utc)
    hace_5min = now - timedelta(minutes=5)
    alertas = await db.alertas.find({'fecha_alerta': {'$gte': hace_5min.isoformat()}}, {'_id': 0}).to_list(1000)
    enviadas = sum(1 for a in alertas if a.get('enviado'))
    sin_tel = sum(1 for a in alertas if not a.get('telefono'))
    errores = max(len(alertas) - enviadas - sin_tel, 0)
    return AlertaResult(enviadas=enviadas, errores=errores, sin_telefono=sin_tel, detalle=alertas)

# ============ EXPORTACION EXCEL ============

def _estilo_encabezado(ws, fila, columnas, color_hex="1E40AF"):
    fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")
    font = Font(color="FFFFFF", bold=True)
    for col_letter in columnas:
        cell = ws[f"{col_letter}{fila}"]
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")

def _autoajustar_columnas(ws):
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

@api_router.get("/exportar/socios")
async def exportar_socios_excel(current_user: dict = Depends(get_current_user)):
    socios = await db.socios.find({}, {'_id': 0}).sort('fecha_registro', -1).to_list(10000)
    pagos = await db.pagos.find({}, {'_id': 0}).sort('fecha_pago', -1).to_list(100000)
    pagos_por_socio = {}
    for p in pagos:
        pagos_por_socio.setdefault(p['socio_id'], []).append(p)

    wb = openpyxl.Workbook()

    # Hoja 1: Socios
    ws1 = wb.active
    ws1.title = "Socios"
    ws1.row_dimensions[1].height = 20
    enc1 = ["ID Socio", "Nombre", "Email", "Teléfono", "Dirección", "Estado", "Fecha Registro", "Vencimiento"]
    ws1.append(enc1)
    _estilo_encabezado(ws1, 1, [chr(65 + i) for i in range(len(enc1))])
    for s in socios:
        ws1.append([s.get('socio_id', ''), s.get('nombre', ''), s.get('email', '') or '', s.get('telefono', '') or '', s.get('direccion', '') or '', 'Activo' if s.get('estado') == 'activo' else 'Vencido', s.get('fecha_registro', '')[:10] if s.get('fecha_registro') else '', s.get('fecha_vencimiento', '')[:10] if s.get('fecha_vencimiento') else 'Sin pagos'])
    _autoajustar_columnas(ws1)

    # Hoja 2: Pagos
    ws2 = wb.create_sheet("Pagos")
    ws2.row_dimensions[1].height = 20
    enc2 = ["ID Socio", "Nombre Socio", "Plan", "Monto", "Método Pago", "Fecha Pago", "Fecha Vencimiento"]
    ws2.append(enc2)
    _estilo_encabezado(ws2, 1, [chr(65 + i) for i in range(len(enc2))], color_hex="065F46")
    socio_nombres = {s['socio_id']: s['nombre'] for s in socios}
    for p in pagos:
        ws2.append([p.get('socio_id', ''), socio_nombres.get(p['socio_id'], 'Desconocido'), p.get('tipo_plan', '').capitalize(), p.get('monto', 0), p.get('metodo_pago', ''), p.get('fecha_pago', '')[:10] if p.get('fecha_pago') else '', p.get('fecha_vencimiento', '')[:10] if p.get('fecha_vencimiento') else ''])
    _autoajustar_columnas(ws2)

    # Hoja 3: Historial por socio
    ws3 = wb.create_sheet("Historial por Socio")
    fila = 1
    for s in socios:
        ws3.cell(row=fila, column=1, value=f"{s['socio_id']} — {s['nombre']}").font = Font(bold=True, size=12)
        fila += 1
        cols_h = ["Plan", "Monto", "Método", "Fecha Pago", "Vencimiento"]
        for i, col in enumerate(cols_h, 1):
            cell = ws3.cell(row=fila, column=i, value=col)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4B5563", end_color="4B5563", fill_type="solid")
        fila += 1
        historial = pagos_por_socio.get(s['socio_id'], [])
        if historial:
            for p in historial:
                ws3.cell(row=fila, column=1, value=p.get('tipo_plan', '').capitalize())
                ws3.cell(row=fila, column=2, value=p.get('monto', 0))
                ws3.cell(row=fila, column=3, value=p.get('metodo_pago', ''))
                ws3.cell(row=fila, column=4, value=p.get('fecha_pago', '')[:10])
                ws3.cell(row=fila, column=5, value=p.get('fecha_vencimiento', '')[:10])
                fila += 1
        else:
            c = ws3.cell(row=fila, column=1, value="Sin pagos registrados")
            c.font = Font(italic=True, color="9CA3AF")
            fila += 1
        fila += 1
    for col in ws3.columns:
        ws3.column_dimensions[col[0].column_letter].width = 20

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    fecha_hoy = datetime.now().strftime('%Y%m%d')
    return StreamingResponse(buffer, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f"attachment; filename=gimnasio_socios_{fecha_hoy}.xlsx"})

@api_router.get("/exportar/pagos")
async def exportar_pagos_excel(current_user: dict = Depends(get_current_user)):
    pagos = await db.pagos.find({}, {'_id': 0}).sort('fecha_pago', -1).to_list(100000)
    socios = await db.socios.find({}, {'_id': 0}).to_list(10000)
    socio_nombres = {s['socio_id']: s['nombre'] for s in socios}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pagos"
    ws.row_dimensions[1].height = 20
    enc = ["ID Socio", "Nombre", "Plan", "Monto", "Método Pago", "Fecha Pago", "Vencimiento"]
    ws.append(enc)
    _estilo_encabezado(ws, 1, [chr(65 + i) for i in range(len(enc))], color_hex="065F46")
    total = 0
    for p in pagos:
        monto = p.get('monto', 0)
        total += monto
        ws.append([p.get('socio_id', ''), socio_nombres.get(p['socio_id'], 'Desconocido'), p.get('tipo_plan', '').capitalize(), monto, p.get('metodo_pago', ''), p.get('fecha_pago', '')[:10] if p.get('fecha_pago') else '', p.get('fecha_vencimiento', '')[:10] if p.get('fecha_vencimiento') else ''])
    fila_total = len(pagos) + 2
    ws.cell(row=fila_total, column=3, value="TOTAL").font = Font(bold=True)
    ws.cell(row=fila_total, column=4, value=total).font = Font(bold=True)
    _autoajustar_columnas(ws)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    fecha_hoy = datetime.now().strftime('%Y%m%d')
    return StreamingResponse(buffer, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f"attachment; filename=gimnasio_pagos_{fecha_hoy}.xlsx"})

# ============ APP STARTUP / SHUTDOWN ============

app.include_router(api_router)
app.add_middleware(CORSMiddleware, allow_credentials=True, allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','), allow_methods=["*"], allow_headers=["*"])

@app.on_event("startup")
async def startup():
    scheduler.add_job(ejecutar_alertas_diarias, CronTrigger(hour=9, minute=0), id="alertas_diarias", replace_existing=True)
    scheduler.start()
    logger.info("Scheduler iniciado — alertas diarias a las 9:00 AM UTC")

@app.on_event("shutdown")
async def shutdown_db_client():
    scheduler.shutdown()
    client.close()
