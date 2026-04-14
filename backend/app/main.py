from fastapi import FastAPI, APIRouter, HTTPException, Depends, status, Query
from fastapi.responses import StreamingResponse
from starlette.middleware.cors import CORSMiddleware
from pathlib import Path
from typing import List, Optional
import uuid
import calendar
import logging
from datetime import datetime, timezone, timedelta
import urllib.parse
import io

from apscheduler.schedulers.asyncio import AsyncIOScheduler
from apscheduler.triggers.cron import CronTrigger
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# Importar módulos refactorizados
from app.core.security import get_current_user, create_token, db
from app.models.user import UserRegister, UserLogin, User, TokenResponse
from app.models.socio import SocioCreate, Socio
from app.models.pago import PagoCreate, Pago
from app.models.dashboard import DashboardStats, IngresoPorDia, Alerta, AlertaEnviada, ConfigMensaje
from app.models.plan import Plan, PlanCreate, PlanUpdate
from app.services.auth_service import register_user, login_user
from app.utils.helpers import (
    calcular_fecha_vencimiento, parse_iso_datetime, get_next_sequence,
    initialize_socio_counter, actualizar_estado_socio
)

# Configuración
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# Aplicación FastAPI
app = FastAPI(title="Gym Manager API", version="1.0.0")
api_router = APIRouter(prefix="/api")
scheduler = AsyncIOScheduler()

# ============ ALERTAS - SCHEDULER ============

async def ejecutar_alertas_diarias():
    """Generar alertas diarias para socios"""
    logger.info("Generando alertas diarias...")
    now = datetime.now(timezone.utc)
    socios = await db.socios.find({}, {'_id': 0}).to_list(10000)

    # Obtener mensajes personalizados
    config_mensajes = await db.config_mensajes.find({}, {'_id': 0}).to_list(100)
    mensajes = {c['tipo']: c['mensaje'] for c in config_mensajes}

    # Limpiar alertas anteriores
    await db.alertas.delete_many({})

    for socio in socios:
        telefono = socio.get('telefono', '').strip()
        if not telefono:
            continue

        alertas_socio = []

        # Vencimientos
        if socio.get('fecha_vencimiento'):
            fecha_venc = datetime.fromisoformat(socio['fecha_vencimiento'])
            dias = (fecha_venc - now).days

            if dias < 0:
                tipo = 'vencido'
            elif 0 <= dias <= 7:
                tipo = 'proximo'
            else:
                tipo = None

            if tipo:
                ya_enviada = await db.alertas_enviadas.find_one({'socio_id': socio['socio_id'], 'tipo': tipo})
                if not ya_enviada:
                    mensaje_default = {
                        'vencido': f"Hola {socio['nombre']} 👋\nTe recordamos que tu cuota en el gimnasio *venció* el {fecha_venc.strftime('%d/%m/%Y')}.\n¡Acercate a renovarla para seguir entrenando! 💪",
                        'proximo': f"Hola {socio['nombre']} 👋\nTe avisamos que tu cuota del gimnasio vence en *{dias} días* ({fecha_venc.strftime('%d/%m/%Y')}).\n¡No te olvides de renovarla! 💪"
                    }[tipo]
                    mensaje = mensajes.get(tipo, mensaje_default)
                    mensaje = mensaje.replace("{nombre}", socio['nombre']).replace("{fecha}", fecha_venc.strftime('%d/%m/%Y')).replace("{dias}", str(dias))
                    prioridad = {'vencido': 1, 'proximo': 2 if dias <= 1 else 3}[tipo]
                    alertas_socio.append({
                        'tipo': tipo,
                        'dias_restantes': dias,
                        'mensaje': mensaje,
                        'prioridad': prioridad
                    })

        # Inactividad
        ultimo_pago = await db.pagos.find_one({'socio_id': socio['socio_id']}, {'_id': 0}, sort=[('fecha_pago', -1)])
        if ultimo_pago:
            fecha_ult_pago = datetime.fromisoformat(ultimo_pago['fecha_pago'])
            dias_sin_pago = (now - fecha_ult_pago).days
            if dias_sin_pago > 30:
                tipo = 'inactivo'
                ya_enviada = await db.alertas_enviadas.find_one({'socio_id': socio['socio_id'], 'tipo': tipo})
                if not ya_enviada:
                    mensaje_default = f"Hola {socio['nombre']} 👋\nHace {dias_sin_pago} días que no nos visitas al gimnasio.\n¡Te esperamos de vuelta! 💪"
                    mensaje = mensajes.get(tipo, mensaje_default)
                    mensaje = mensaje.replace("{nombre}", socio['nombre']).replace("{dias}", str(dias_sin_pago))
                    alertas_socio.append({
                        'tipo': tipo,
                        'dias_restantes': None,
                        'mensaje': mensaje,
                        'prioridad': 4
                    })

        # Guardar alertas
        for alerta in alertas_socio:
            whatsapp_link = f"https://wa.me/{telefono.replace('+', '')}?text={urllib.parse.quote(alerta['mensaje'])}"
            alerta_doc = {
                'id': str(uuid.uuid4()),
                'socio_id': socio['socio_id'],
                'nombre': socio['nombre'],
                'tipo': alerta['tipo'],
                'dias_restantes': alerta.get('dias_restantes'),
                'fecha_vencimiento': socio.get('fecha_vencimiento'),
                'telefono': telefono,
                'mensaje': alerta['mensaje'],
                'whatsapp_link': whatsapp_link,
                'fecha_alerta': now.isoformat(),
                'prioridad': alerta['prioridad']
            }
            await db.alertas.insert_one(alerta_doc)

    logger.info("Alertas generadas exitosamente")

# ============ AUTH ROUTES ============

@api_router.post("/auth/register", response_model=TokenResponse)
async def register(user_data: UserRegister):
    """Registrar nuevo usuario"""
    return await register_user(user_data)

@api_router.post("/auth/login", response_model=TokenResponse)
async def login(credentials: UserLogin):
    """Autenticar usuario"""
    return await login_user(credentials)

@api_router.get("/auth/me", response_model=User)
async def get_me(current_user: dict = Depends(get_current_user)):
    """Obtener datos del usuario autenticado"""
    return User(**current_user)

# ============ SOCIOS ROUTES ============

@api_router.post("/socios", response_model=Socio)
async def crear_socio(socio_data: SocioCreate, current_user: dict = Depends(get_current_user)):
    """Crear nuevo socio"""
    next_number = await get_next_sequence('socio_id')
    socio_id = f"GYM-{next_number:04d}"
    socio_doc = {
        'id': str(uuid.uuid4()),
        'socio_id': socio_id,
        'nombre': socio_data.nombre,
        'apellido': socio_data.apellido,
        'email': socio_data.email,
        'telefono': socio_data.telefono,
        'direccion': socio_data.direccion,
        'dni': socio_data.dni,
        'fecha_registro': datetime.now(timezone.utc).isoformat(),
        'estado': 'vencido',
        'fecha_vencimiento': None
    }
    await db.socios.insert_one(socio_doc)
    socio_doc.pop('_id', None)
    return Socio(**socio_doc)

@api_router.get("/socios", response_model=List[Socio])
async def listar_socios(current_user: dict = Depends(get_current_user)):
    """Listar todos los socios"""
    socios = await db.socios.find({}, {'_id': 0}).sort('socio_id', 1).to_list(1000)
    for socio in socios:
        socio.setdefault('apellido', '')
        socio.setdefault('dni', '')
    return [Socio(**s) for s in socios]

@api_router.get("/socios/{socio_id}", response_model=Socio)
async def obtener_socio(socio_id: str, current_user: dict = Depends(get_current_user)):
    """Obtener detalle de un socio"""
    socio = await db.socios.find_one({'socio_id': socio_id}, {'_id': 0})
    if not socio:
        raise HTTPException(status_code=404, detail="Socio no encontrado")
    socio.setdefault('apellido', '')
    socio.setdefault('dni', '')
    return Socio(**socio)

@api_router.put("/socios/{socio_id}", response_model=Socio)
async def actualizar_socio(socio_id: str, socio_data: SocioCreate, current_user: dict = Depends(get_current_user)):
    """Actualizar datos de un socio"""
    result = await db.socios.update_one({'socio_id': socio_id}, {'$set': socio_data.model_dump()})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Socio no encontrado")
    socio = await db.socios.find_one({'socio_id': socio_id}, {'_id': 0})
    return Socio(**socio)

@api_router.delete("/socios/{socio_id}")
async def eliminar_socio(socio_id: str, current_user: dict = Depends(get_current_user)):
    """Eliminar un socio"""
    result = await db.socios.delete_one({'socio_id': socio_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Socio no encontrado")
    await db.pagos.delete_many({'socio_id': socio_id})
    return {'message': 'Socio eliminado correctamente'}

# ============ PAGOS ROUTES ============

@api_router.post("/pagos", response_model=Pago)
async def registrar_pago(pago_data: PagoCreate, current_user: dict = Depends(get_current_user)):
    """Registrar nuevo pago"""
    socio = await db.socios.find_one({'socio_id': pago_data.socio_id}, {'_id': 0})
    if not socio:
        raise HTTPException(status_code=404, detail="Socio no encontrado")
    fecha_pago = datetime.now(timezone.utc)
    fecha_vencimiento = calcular_fecha_vencimiento(pago_data.tipo_plan, fecha_pago)
    pago_doc = {
        'id': str(uuid.uuid4()),
        'socio_id': pago_data.socio_id,
        'monto': pago_data.monto,
        'tipo_plan': pago_data.tipo_plan,
        'metodo_pago': pago_data.metodo_pago,
        'fecha_pago': fecha_pago.isoformat(),
        'fecha_vencimiento': fecha_vencimiento.isoformat()
    }
    await db.pagos.insert_one(pago_doc)
    await actualizar_estado_socio(pago_data.socio_id)
    pago_doc.pop('_id', None)
    return Pago(**pago_doc)

@api_router.get("/pagos", response_model=List[Pago])
async def listar_pagos(current_user: dict = Depends(get_current_user)):
    """Listar todos los pagos"""
    pagos = await db.pagos.find({}, {'_id': 0}).sort('fecha_pago', -1).to_list(1000)
    return [Pago(**p) for p in pagos]

@api_router.get("/pagos/socio/{socio_id}", response_model=List[Pago])
async def obtener_pagos_socio(socio_id: str, current_user: dict = Depends(get_current_user)):
    """Obtener pagos de un socio"""
    pagos = await db.pagos.find({'socio_id': socio_id}, {'_id': 0}).sort('fecha_pago', -1).to_list(1000)
    return [Pago(**p) for p in pagos]

# ============ DASHBOARD ROUTES ============

@api_router.get("/dashboard/stats", response_model=DashboardStats)
async def obtener_stats(
    month: Optional[int] = Query(None, ge=1, le=12),
    year: Optional[int] = Query(None, ge=2000),
    current_user: dict = Depends(get_current_user)
):
    """Obtener estadísticas del dashboard"""
    socios = await db.socios.find({}, {'_id': 0}).to_list(10000)
    for socio in socios:
        await actualizar_estado_socio(socio['socio_id'])
    socios = await db.socios.find({}, {'_id': 0}).to_list(10000)
    
    total_socios = len(socios)
    socios_activos = len([s for s in socios if s.get('estado') == 'activo'])
    socios_vencidos = len([s for s in socios if s.get('estado') == 'vencido'])
    
    now = datetime.now(timezone.utc)
    if month is None:
        month = now.month
    if year is None:
        year = now.year
    
    pagos = await db.pagos.find({}, {'_id': 0}).to_list(10000)
    pagos_mes = []
    for pago in pagos:
        try:
            fecha_pago = parse_iso_datetime(pago['fecha_pago'])
        except Exception:
            continue
        if fecha_pago.year == year and fecha_pago.month == month:
            pagos_mes.append(pago)
    
    ingresos_mes = sum(p['monto'] for p in pagos_mes)

    ingresos_por_dia_dict = {}
    for pago in pagos_mes:
        try:
            fecha_pago = parse_iso_datetime(pago['fecha_pago'])
        except Exception:
            continue
        dia = fecha_pago.day
        if dia not in ingresos_por_dia_dict:
            ingresos_por_dia_dict[dia] = {'ingresos': 0.0, 'pagos': 0}
        ingresos_por_dia_dict[dia]['ingresos'] += pago['monto']
        ingresos_por_dia_dict[dia]['pagos'] += 1

    days_in_month = calendar.monthrange(year, month)[1]
    ingresos_por_dia = []
    for dia in range(1, days_in_month + 1):
        ingresos_por_dia.append(IngresoPorDia(
            dia=dia,
            fecha=f"{year}-{month:02d}-{dia:02d}",
            ingresos=round(ingresos_por_dia_dict.get(dia, {'ingresos': 0.0})['ingresos'], 2),
            pagos=ingresos_por_dia_dict.get(dia, {'pagos': 0})['pagos']
        ))

    proximos = []
    for socio in socios:
        if socio.get('fecha_vencimiento'):
            fecha_venc = parse_iso_datetime(socio['fecha_vencimiento'])
            dias_restantes = (fecha_venc - now).days
            if 0 <= dias_restantes <= 7:
                proximos.append({
                    'socio_id': socio['socio_id'],
                    'nombre': socio['nombre'],
                    'fecha_vencimiento': socio['fecha_vencimiento'],
                    'dias_restantes': dias_restantes
                })
    proximos.sort(key=lambda x: x['dias_restantes'])
    
    return DashboardStats(
        total_socios=total_socios,
        socios_activos=socios_activos,
        socios_vencidos=socios_vencidos,
        ingresos_mes=ingresos_mes,
        mes=month,
        anio=year,
        ingresos_por_dia=ingresos_por_dia,
        proximos_vencimientos=proximos[:10]
    )

# ============ ALERTAS ROUTES ============

@api_router.get("/alertas", response_model=List[Alerta])
async def obtener_alertas(current_user: dict = Depends(get_current_user)):
    """Obtener alertas activas"""
    alertas = await db.alertas.find({}, {'_id': 0}).sort('prioridad', 1).to_list(1000)
    return [Alerta(**a) for a in alertas]

@api_router.get("/alertas/enviadas")
async def obtener_alertas_enviadas(current_user: dict = Depends(get_current_user)):
    """Obtener historial de alertas enviadas"""
    pipeline = [
        {
            "$lookup": {
                "from": "socios",
                "localField": "socio_id",
                "foreignField": "socio_id",
                "as": "socio"
            }
        },
        {
            "$unwind": "$socio"
        },
        {
            "$project": {
                "_id": 0,
                "id": 1,
                "socio_id": 1,
                "nombre": "$socio.nombre",
                "apellido": "$socio.apellido",
                "tipo": 1,
                "fecha_envio": 1
            }
        },
        {
            "$sort": {"fecha_envio": -1}
        }
    ]
    historial = await db.alertas_enviadas.aggregate(pipeline).to_list(1000)
    return historial

@api_router.post("/alertas/{alerta_id}/enviar")
async def enviar_alerta(alerta_id: str, current_user: dict = Depends(get_current_user)):
    """Marcar alerta como enviada"""
    alerta = await db.alertas.find_one({'id': alerta_id}, {'_id': 0})
    if not alerta:
        raise HTTPException(status_code=404, detail="Alerta no encontrada")

    enviada_doc = {
        'id': str(uuid.uuid4()),
        'socio_id': alerta['socio_id'],
        'tipo': alerta['tipo'],
        'fecha_envio': datetime.now(timezone.utc).isoformat()
    }
    await db.alertas_enviadas.insert_one(enviada_doc)
    await db.alertas.delete_one({'id': alerta_id})

    return {"message": "Alerta marcada como enviada"}

@api_router.post("/alertas/generar")
async def generar_alertas(current_user: dict = Depends(get_current_user)):
    """Generar alertas manualmente"""
    await ejecutar_alertas_diarias()
    alertas = await db.alertas.find({}, {'_id': 0}).to_list(1000)
    return {"alertas_generadas": len(alertas)}

# ============ CONFIG MENSAJES ROUTES ============

@api_router.get("/config/mensajes", response_model=List[ConfigMensaje])
async def obtener_config_mensajes(current_user: dict = Depends(get_current_user)):
    """Obtener configuración de mensajes"""
    mensajes = await db.config_mensajes.find({}, {'_id': 0}).to_list(100)
    if not mensajes:
        defaults = [
            {"tipo": "vencido", "mensaje": "Hola {nombre} 👋\nTe recordamos que tu cuota en el gimnasio *venció* el {fecha}.\n¡Acercate a renovarla para seguir entrenando! 💪"},
            {"tipo": "proximo", "mensaje": "Hola {nombre} 👋\nTe avisamos que tu cuota del gimnasio vence en *{dias} días* ({fecha}).\n¡No te olvides de renovarla! 💪"},
            {"tipo": "inactivo", "mensaje": "Hola {nombre} 👋\nHace {dias} días que no nos visitas al gimnasio.\n¡Te esperamos de vuelta! 💪"}
        ]
        for d in defaults:
            await db.config_mensajes.insert_one(d)
        mensajes = defaults
    return [ConfigMensaje(**m) for m in mensajes]

@api_router.put("/config/mensajes")
async def actualizar_config_mensajes(configs: List[ConfigMensaje], current_user: dict = Depends(get_current_user)):
    """Actualizar configuración de mensajes"""
    await db.config_mensajes.delete_many({})
    for config in configs:
        await db.config_mensajes.insert_one(config.model_dump())
    return {"message": "Mensajes actualizados"}

# ============ PLANES ROUTES ============

@api_router.get("/planes", response_model=List[Plan])
async def obtener_planes(current_user: dict = Depends(get_current_user)):
    """Obtener planes disponibles"""
    planes = await db.planes.find({}, {'_id': 0}).to_list(100)
    if not planes:
        defaults = [
            {"id": str(uuid.uuid4()), "nombre": "Mensual", "dias": 30, "precio": 50.0},
            {"id": str(uuid.uuid4()), "nombre": "Trimestral", "dias": 90, "precio": 130.0},
            {"id": str(uuid.uuid4()), "nombre": "Semestral", "dias": 180, "precio": 240.0},
            {"id": str(uuid.uuid4()), "nombre": "Anual", "dias": 365, "precio": 450.0}
        ]
        for d in defaults:
            await db.planes.insert_one(d)
        planes = defaults
    return [Plan(**p) for p in planes]

@api_router.post("/planes")
async def crear_plan(plan: PlanCreate, current_user: dict = Depends(get_current_user)):
    """Crear nuevo plan"""
    plan_doc = {
        "id": str(uuid.uuid4()),
        "nombre": plan.nombre,
        "dias": plan.dias,
        "precio": plan.precio
    }
    await db.planes.insert_one(plan_doc)
    plan_doc.pop('_id', None)
    return plan_doc

@api_router.put("/planes/{plan_id}")
async def actualizar_plan(plan_id: str, plan: PlanUpdate, current_user: dict = Depends(get_current_user)):
    """Actualizar un plan"""
    result = await db.planes.update_one(
        {'id': plan_id},
        {'$set': {'nombre': plan.nombre, 'dias': plan.dias, 'precio': plan.precio}}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Plan no encontrado")
    return {"message": "Plan actualizado"}

@api_router.delete("/planes/{plan_id}")
async def eliminar_plan(plan_id: str, current_user: dict = Depends(get_current_user)):
    """Eliminar un plan"""
    result = await db.planes.delete_one({'id': plan_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Plan no encontrado")
    return {"message": "Plan eliminado"}

# ============ EXPORTACION EXCEL ============

def _estilo_encabezado(ws, fila, columnas, color_hex="1E40AF"):
    """Aplicar estilo a encabezados"""
    fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")
    font = Font(color="FFFFFF", bold=True)
    for col_letter in columnas:
        cell = ws[f"{col_letter}{fila}"]
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")

def _autoajustar_columnas(ws):
    """Autoajustar ancho de columnas"""
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

@api_router.get("/exportar/socios")
async def exportar_socios_excel(current_user: dict = Depends(get_current_user)):
    """Exportar socios a Excel"""
    socios = await db.socios.find({}, {'_id': 0}).sort('fecha_registro', -1).to_list(10000)
    pagos = await db.pagos.find({}, {'_id': 0}).sort('fecha_pago', -1).to_list(100000)
    pagos_por_socio = {}
    for p in pagos:
        pagos_por_socio.setdefault(p['socio_id'], []).append(p)

    wb = openpyxl.Workbook()

    # Hoja 1: Socios
    ws1 = wb.active
    ws1.title = "Socios"
    ws1.row_dimensions[1].height = 20
    enc1 = ["ID Socio", "Nombre", "Apellido", "DNI", "Email", "Teléfono", "Dirección", "Estado", "Fecha Registro", "Vencimiento"]
    ws1.append(enc1)
    _estilo_encabezado(ws1, 1, [chr(65 + i) for i in range(len(enc1))])
    for s in socios:
        ws1.append([
            s.get('socio_id', ''),
            s.get('nombre', ''),
            s.get('apellido', ''),
            s.get('dni', ''),
            s.get('email', '') or '',
            s.get('telefono', '') or '',
            s.get('direccion', '') or '',
            'Activo' if s.get('estado') == 'activo' else 'Vencido',
            s.get('fecha_registro', '')[:10] if s.get('fecha_registro') else '',
            s.get('fecha_vencimiento', '')[:10] if s.get('fecha_vencimiento') else 'Sin pagos'
        ])
    _autoajustar_columnas(ws1)

    # Hoja 2: Pagos
    ws2 = wb.create_sheet("Pagos")
    ws2.row_dimensions[1].height = 20
    enc2 = ["ID Socio", "Nombre Socio", "Plan", "Monto", "Método Pago", "Fecha Pago", "Fecha Vencimiento"]
    ws2.append(enc2)
    _estilo_encabezado(ws2, 1, [chr(65 + i) for i in range(len(enc2))], color_hex="065F46")
    socio_nombres = {s['socio_id']: f"{s['nombre']} {s['apellido']}" for s in socios}
    for p in pagos:
        ws2.append([
            p.get('socio_id', ''),
            socio_nombres.get(p['socio_id'], 'Desconocido'),
            p.get('tipo_plan', '').capitalize(),
            p.get('monto', 0),
            p.get('metodo_pago', ''),
            p.get('fecha_pago', '')[:10] if p.get('fecha_pago') else '',
            p.get('fecha_vencimiento', '')[:10] if p.get('fecha_vencimiento') else ''
        ])
    _autoajustar_columnas(ws2)

    # Hoja 3: Historial por socio
    ws3 = wb.create_sheet("Historial por Socio")
    fila = 1
    for s in socios:
        ws3.cell(row=fila, column=1, value=f"{s['socio_id']} — {s['nombre']} {s['apellido']}").font = Font(bold=True, size=12)
        fila += 1
        cols_h = ["Plan", "Monto", "Método", "Fecha Pago", "Vencimiento"]
        for i, col in enumerate(cols_h, 1):
            cell = ws3.cell(row=fila, column=i, value=col)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4B5563", end_color="4B5563", fill_type="solid")
        fila += 1
        historial = pagos_por_socio.get(s['socio_id'], [])
        if historial:
            for p in historial:
                ws3.cell(row=fila, column=1, value=p.get('tipo_plan', '').capitalize())
                ws3.cell(row=fila, column=2, value=p.get('monto', 0))
                ws3.cell(row=fila, column=3, value=p.get('metodo_pago', ''))
                ws3.cell(row=fila, column=4, value=p.get('fecha_pago', '')[:10])
                ws3.cell(row=fila, column=5, value=p.get('fecha_vencimiento', '')[:10])
                fila += 1
        else:
            c = ws3.cell(row=fila, column=1, value="Sin pagos registrados")
            c.font = Font(italic=True, color="9CA3AF")
            fila += 1
        fila += 1
    for col in ws3.columns:
        ws3.column_dimensions[col[0].column_letter].width = 20

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    fecha_hoy = datetime.now().strftime('%Y%m%d')
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=gimnasio_socios_{fecha_hoy}.xlsx"}
    )

@api_router.get("/exportar/pagos")
async def exportar_pagos_excel(current_user: dict = Depends(get_current_user)):
    """Exportar pagos a Excel"""
    pagos = await db.pagos.find({}, {'_id': 0}).sort('fecha_pago', -1).to_list(100000)
    socios = await db.socios.find({}, {'_id': 0}).to_list(10000)
    socio_nombres = {s['socio_id']: s['nombre'] for s in socios}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pagos"
    ws.row_dimensions[1].height = 20
    enc = ["ID Socio", "Nombre", "Plan", "Monto", "Método Pago", "Fecha Pago", "Vencimiento"]
    ws.append(enc)
    _estilo_encabezado(ws, 1, [chr(65 + i) for i in range(len(enc))], color_hex="065F46")
    total = 0
    for p in pagos:
        monto = p.get('monto', 0)
        total += monto
        ws.append([
            p.get('socio_id', ''),
            socio_nombres.get(p['socio_id'], 'Desconocido'),
            p.get('tipo_plan', '').capitalize(),
            monto,
            p.get('metodo_pago', ''),
            p.get('fecha_pago', '')[:10] if p.get('fecha_pago') else '',
            p.get('fecha_vencimiento', '')[:10] if p.get('fecha_vencimiento') else ''
        ])
    fila_total = len(pagos) + 2
    ws.cell(row=fila_total, column=3, value="TOTAL").font = Font(bold=True)
    ws.cell(row=fila_total, column=4, value=total).font = Font(bold=True)
    _autoajustar_columnas(ws)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    fecha_hoy = datetime.now().strftime('%Y%m%d')
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=gimnasio_pagos_{fecha_hoy}.xlsx"}
    )

# ============ APP STARTUP / SHUTDOWN ============

app.include_router(api_router)
import os
from dotenv import load_dotenv
load_dotenv(Path(__file__).parent.parent / '.env')

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"]
)

@app.on_event("startup")
async def startup():
    """Inicializar aplicación"""
    await db.socios.create_index('socio_id', unique=True)
    await db.counters.create_index('name', unique=True)
    await initialize_socio_counter()
    scheduler.add_job(
        ejecutar_alertas_diarias,
        CronTrigger(hour=9, minute=0),
        id="alertas_diarias",
        replace_existing=True
    )
    scheduler.start()
    logger.info("✅ Servidor iniciado — alertas diarias a las 9:00 AM UTC")

@app.on_event("shutdown")
async def shutdown_db_client():
    """Cerrar conexiones al apagar"""
    scheduler.shutdown()
    db.client.close()
    logger.info("✅ Servidor apagado correctamente")
